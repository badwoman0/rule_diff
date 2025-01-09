# This module will handle file uploads and comparisons

def handle_upload(file):
    # Logic to handle uploaded file
    pass


def compare_files(file1, file2):
    # Logic to compare two files and generate a CSV of differences
    pass


def parse_conf(file_path):
    parsed_data = {}

    with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
        for line_number, line in enumerate(file, start=1):
            try:
                # 去除行首尾的空白字符
                line = line.strip()
                if not line:
                    continue

                # 分割行内容
                parts = line.split()
                if len(parts) < 6:
                    continue

                # 提取信息
                config_name = parts[0].split(':')[1]
                rule_id = parts[1]
                rule_level = parts[2]
                default_enabled = parts[3]
                log_record = parts[4]
                block_mode = parts[5]

                # 构建字典结构
                if config_name not in parsed_data:
                    parsed_data[config_name] = []

                parsed_data[config_name].append({
                    "规则id": rule_id,
                    "规则等级": rule_level,
                    "是否默认启用": default_enabled,
                    "日志记录": log_record,
                    "阻断方式": block_mode
                })
            except Exception as e:
                raise ValueError(f"Error parsing line {line_number}: {e}")

    return parsed_data


def compare_configs(default_conf_path, customer_conf_path):
    # 解析两个配置文件
    default_data = parse_conf(default_conf_path)
    customer_data = parse_conf(customer_conf_path)

    # 存储差异的列表
    differences = []

    # 获取默认配置中的qw_new配置
    qw_new_config = default_data.get("qw_new", [])

    # 读取WAF_rules.xlsx以获取规则名称
    rules_df = pd.read_excel('/Users/badwoman/waf_new/rule_diff/config/WAF_rules.xlsx')
    rules_dict = dict(zip(rules_df['历史规则ID映射'], rules_df['规则名称']))

    # 遍历客户配置中的配置名
    for config_name, rules in customer_data.items():
        for rule in rules:
            rule_id = rule["规则id"]
            # 在qw_new配置中查找相同规则ID的规则
            for default_rule in qw_new_config:
                if default_rule["规则id"] == rule_id:
                    # 比较是否默认启用和阻断方式
                    if (default_rule["是否默认启用"] != rule["是否默认启用"] or
                        default_rule["阻断方式"] != rule["阻断方式"]):
                        differences.append({
                            "配置名": config_name,
                            "规则id": hex(int(rule_id)),
                            "规则名称": rules_dict.get(hex(int(rule_id)), ""),
                            "默认配置是否默认启用": default_rule["是否默认启用"],
                            "现有配置是否默认启用": rule["是否默认启用"],
                            "默认配置阻断方式": default_rule["阻断方式"],
                            "现有配置阻断方式": rule["阻断方式"]
                        })

    return differences


import pandas as pd
from openpyxl.styles import PatternFill

def write_differences_to_xlsx(differences, xlsx_path):
    df = pd.DataFrame(differences)
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Differences')
        worksheet = writer.sheets['Differences']
        # 设置颜色和格式
        for index, row in df.iterrows():
            # 比较“是否默认启用”
            default_enable_cell = worksheet.cell(row=index+2, column=df.columns.get_loc("默认配置是否默认启用")+1)
            current_enable_cell = worksheet.cell(row=index+2, column=df.columns.get_loc("现有配置是否默认启用")+1)
            if row["默认配置是否默认启用"] != row["现有配置是否默认启用"]:
                default_enable_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                current_enable_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            # 比较“阻断方式”
            default_block_cell = worksheet.cell(row=index+2, column=df.columns.get_loc("默认配置阻断方式")+1)
            current_block_cell = worksheet.cell(row=index+2, column=df.columns.get_loc("现有配置阻断方式")+1)
            if row["默认配置阻断方式"] != row["现有配置阻断方式"]:
                default_block_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                current_block_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# 示例调用
# file_path = 'path/to/your/config.conf'
# parsed_dict = parse_conf(file_path)
# print(parsed_dict)

# differences = compare_configs('path/to/default.conf', 'path/to/customer.conf')
# print(differences)
# write_differences_to_xlsx(differences, 'path/to/differences.xlsx')
